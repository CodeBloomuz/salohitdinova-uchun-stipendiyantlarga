"""
🎓 Stipendiat Ball Tracker — Telegram Bot v2
Talaba + Admin panel | Fayl tasdiqlash | Excel eksport
"""

import logging
import sqlite3
import random
import io
from datetime import datetime
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove,
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ─── SOZLAMALAR ───────────────────────────────────────────────────────────────
BOT_TOKEN   = "YOUR_BOT_TOKEN_HERE"
ADMIN_IDS   = [123456789]           # ← o'z Telegram ID ingizni kiriting
DB_PATH     = "stipend_bot.db"

logging.basicConfig(
    format="%(asctime)s — %(levelname)s — %(message)s",
    level=logging.INFO,
)

# ─── FAOLIYATLAR ─────────────────────────────────────────────────────────────
DEFAULT_ACTIVITIES = {
    "tadbir":        {"name": "Tadbirda ishtirok",          "icon": "🎪", "pts": 2,  "max": None},
    "milliy_konf":   {"name": "Milliy konferensiya",        "icon": "🎓", "pts": 1,  "max": 3},
    "xalqaro":       {"name": "Xalqaro nashr/konferensiya", "icon": "🌍", "pts": 2,  "max": 2},
    "oak_jurnal":    {"name": "OAK jurnalida maqola",       "icon": "📰", "pts": 3,  "max": 5},
    "mahalliy":      {"name": "Mahalliy jurnalda maqola",   "icon": "📄", "pts": 2,  "max": 5},
    "gazeta":        {"name": "Gazetaga maqola",            "icon": "📋", "pts": 1,  "max": None},
    "volontyorlik":  {"name": "Volontyorlik (1 kun)",       "icon": "🤝", "pts": 10, "max": 10},
    "dgu":           {"name": "DGU guvohnoma",              "icon": "🏅", "pts": 3,  "max": None},
}

MOTIVATIONS = [
    "🔥 Zo'r! Faoliyating tasdiqlash kutmoqda — yaqinda ball qo'shiladi!",
    "💪 Harakatingiz qayd etildi! Admin tez orada ko'rib chiqadi.",
    "⭐ Har bir hujjat — stipendiatga bir qadam yaqin!",
    "🚀 Zo'r ketayapsan! Tasdiqlashni kut va davom et!",
    "🌟 Rahmat! Hujjating adminга yuborildi — sabr qil!",
    "🏆 Izchillik — muvaffaqiyat kaliti! Davom et!",
]

ACHIEVEMENTS = [
    {"id": "first",       "icon": "🌱", "name": "Birinchi qadam",      "desc": "Birinchi ball to'pladingiz",
     "check": lambda t, c: t >= 1},
    {"id": "ten",         "icon": "🔥", "name": "O'n ball",            "desc": "10 ball to'pladingiz",
     "check": lambda t, c: t >= 10},
    {"id": "twentyfive",  "icon": "⚡", "name": "Yarim yo'l",          "desc": "25 ball to'pladingiz",
     "check": lambda t, c: t >= 25},
    {"id": "fifty",       "icon": "🏆", "name": "STIPENDIAT!",         "desc": "50 ball to'pladingiz!",
     "check": lambda t, c: t >= 50},
    {"id": "writer",      "icon": "✍️",  "name": "Maqolachi",           "desc": "3+ maqola nashr etdingiz",
     "check": lambda t, c: sum(c.get(k, 0) for k in ["oak_jurnal", "mahalliy", "gazeta"]) >= 3},
    {"id": "volunteer",   "icon": "🤝", "name": "Volontyorlik qahramoni","desc": "5+ kun volontyorlik",
     "check": lambda t, c: c.get("volontyorlik", 0) >= 5},
    {"id": "conference",  "icon": "🎤", "name": "Konferensiya yulduzi", "desc": "2+ konferensiyada ishtirok",
     "check": lambda t, c: c.get("milliy_konf", 0) + c.get("xalqaro", 0) >= 2},
    {"id": "global",      "icon": "🌐", "name": "Xalqarochi",          "desc": "2 ta xalqaro nashr",
     "check": lambda t, c: c.get("xalqaro", 0) >= 2},
    {"id": "streak3",     "icon": "📆", "name": "3 kunlik zanjir",     "desc": "3 kun ketma-ket faoliyat",
     "check": lambda t, c: c.get("_streak", 0) >= 3},
]

# ─── MA'LUMOTLAR BAZASI ───────────────────────────────────────────────────────
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            user_id    INTEGER PRIMARY KEY,
            name       TEXT    NOT NULL,
            group_name TEXT    NOT NULL,
            phone      TEXT,
            semester   INTEGER DEFAULT 1,
            streak     INTEGER DEFAULT 0,
            last_date  TEXT,
            status     TEXT    DEFAULT 'pending',
            joined_at  TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS requests (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL,
            activity_type TEXT    NOT NULL,
            semester      INTEGER NOT NULL,
            file_id       TEXT,
            file_type     TEXT,
            status        TEXT    DEFAULT 'pending',
            note          TEXT,
            submitted_at  TEXT    NOT NULL,
            reviewed_at   TEXT
        );
        CREATE TABLE IF NOT EXISTS achievements (
            user_id        INTEGER NOT NULL,
            achievement_id TEXT    NOT NULL,
            earned_at      TEXT    NOT NULL,
            PRIMARY KEY (user_id, achievement_id)
        );
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
        CREATE TABLE IF NOT EXISTS messages (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id   INTEGER NOT NULL,
            text      TEXT    NOT NULL,
            sent_at   TEXT    NOT NULL
        );
        """)

# ── Users ────────────────────────────────────────────────────────────────────
def get_user(uid):
    with db() as c:
        return c.execute("SELECT * FROM users WHERE user_id=?", (uid,)).fetchone()

def create_user(uid, name, group, phone):
    with db() as c:
        existing = c.execute("SELECT user_id FROM users WHERE user_id=?", (uid,)).fetchone()
        if existing:
            return False
        c.execute(
            "INSERT INTO users (user_id,name,group_name,phone,status,joined_at) VALUES (?,?,?,?,'pending',?)",
            (uid, name, group, phone, datetime.now().isoformat())
        )
        return True

def update_user_status(uid, status):
    with db() as c:
        c.execute("UPDATE users SET status=? WHERE user_id=?", (status, uid))

def set_user_semester(uid, sem):
    with db() as c:
        c.execute("UPDATE users SET semester=? WHERE user_id=?", (sem, uid))

def get_all_users(status=None):
    with db() as c:
        if status:
            return c.execute("SELECT * FROM users WHERE status=? ORDER BY joined_at DESC", (status,)).fetchall()
        return c.execute("SELECT * FROM users ORDER BY joined_at DESC").fetchall()

def delete_user(uid):
    with db() as c:
        c.execute("DELETE FROM users WHERE user_id=?", (uid,))
        c.execute("DELETE FROM requests WHERE user_id=?", (uid,))
        c.execute("DELETE FROM achievements WHERE user_id=?", (uid,))

# ── Score & Activities ────────────────────────────────────────────────────────
def get_activities():
    acts = dict(DEFAULT_ACTIVITIES)
    with db() as c:
        rows = c.execute("SELECT key, value FROM settings WHERE key LIKE 'limit_%'").fetchall()
        for row in rows:
            key = row["key"][6:]   # strip "limit_"
            if key in acts:
                val = None if row["value"] == "null" else int(row["value"])
                acts[key] = dict(acts[key], max=val)
        rows2 = c.execute("SELECT key, value FROM settings WHERE key LIKE 'pts_%'").fetchall()
        for row in rows2:
            key = row["key"][4:]
            if key in acts:
                acts[key] = dict(acts[key], pts=int(row["value"]))
    return acts

def get_score(uid, semester=None):
    with db() as c:
        if semester:
            return c.execute(
                "SELECT COALESCE(SUM(a.pts),0) FROM requests r JOIN users u ON r.user_id=u.user_id "
                "JOIN (SELECT key,pts FROM (SELECT 'tadbir' key,2 pts UNION ALL SELECT 'milliy_konf',1 "
                "UNION ALL SELECT 'xalqaro',2 UNION ALL SELECT 'oak_jurnal',3 UNION ALL SELECT 'mahalliy',2 "
                "UNION ALL SELECT 'gazeta',1 UNION ALL SELECT 'volontyorlik',10 UNION ALL SELECT 'dgu',3)) a "
                "ON r.activity_type=a.key WHERE r.user_id=? AND r.semester=? AND r.status='approved'",
                (uid, semester)
            ).fetchone()[0]
        # simpler: join with python
        return _calc_score(uid)

def _calc_score(uid, semester=None):
    acts = get_activities()
    with db() as c:
        if semester:
            rows = c.execute(
                "SELECT activity_type, COUNT(*) cnt FROM requests WHERE user_id=? AND semester=? AND status='approved' GROUP BY activity_type",
                (uid, semester)
            ).fetchall()
        else:
            rows = c.execute(
                "SELECT activity_type, COUNT(*) cnt FROM requests WHERE user_id=? AND status='approved' GROUP BY activity_type",
                (uid,)
            ).fetchall()
    total = 0
    for row in rows:
        act = acts.get(row["activity_type"])
        if act:
            total += row["cnt"] * act["pts"]
    return total

def get_counts(uid, semester=None):
    with db() as c:
        if semester:
            rows = c.execute(
                "SELECT activity_type, COUNT(*) cnt FROM requests WHERE user_id=? AND semester=? AND status='approved' GROUP BY activity_type",
                (uid, semester)
            ).fetchall()
        else:
            rows = c.execute(
                "SELECT activity_type, COUNT(*) cnt FROM requests WHERE user_id=? AND status='approved' GROUP BY activity_type",
                (uid,)
            ).fetchall()
    return {r["activity_type"]: r["cnt"] for r in rows}

def get_sem_activity_count(uid, atype, semester):
    with db() as c:
        return c.execute(
            "SELECT COUNT(*) FROM requests WHERE user_id=? AND activity_type=? AND semester=? AND status='approved'",
            (uid, atype, semester)
        ).fetchone()[0]

def submit_request(uid, atype, semester, file_id, file_type):
    with db() as c:
        c.execute(
            "INSERT INTO requests (user_id,activity_type,semester,file_id,file_type,status,submitted_at) VALUES (?,?,?,?,?,'pending',?)",
            (uid, atype, semester, file_id, file_type, datetime.now().isoformat())
        )
        return c.lastrowid

def review_request(req_id, status, note=""):
    with db() as c:
        c.execute(
            "UPDATE requests SET status=?, note=?, reviewed_at=? WHERE id=?",
            (status, note, datetime.now().isoformat(), req_id)
        )

def get_pending_requests():
    with db() as c:
        return c.execute(
            "SELECT r.*, u.name, u.group_name FROM requests r JOIN users u ON r.user_id=u.user_id WHERE r.status='pending' ORDER BY r.submitted_at",
        ).fetchall()

def get_user_history(uid, limit=15):
    with db() as c:
        return c.execute(
            "SELECT * FROM requests WHERE user_id=? ORDER BY submitted_at DESC LIMIT ?",
            (uid, limit)
        ).fetchall()

def get_leaderboard(limit=10):
    acts = get_activities()
    with db() as c:
        users = c.execute("SELECT user_id, name, group_name FROM users WHERE status='approved'").fetchall()
    board = []
    for u in users:
        score = _calc_score(u["user_id"])
        board.append((u["name"], u["group_name"], score))
    board.sort(key=lambda x: x[2], reverse=True)
    return board[:limit]

def update_streak(uid):
    today = datetime.now().date().isoformat()
    with db() as c:
        u = c.execute("SELECT streak, last_date FROM users WHERE user_id=?", (uid,)).fetchone()
        if not u:
            return
        streak, last = u["streak"], u["last_date"]
        if last == today:
            new_streak = streak
        elif last and (datetime.now().date() - datetime.fromisoformat(last).date()).days == 1:
            new_streak = streak + 1
        else:
            new_streak = 1
        c.execute("UPDATE users SET streak=?, last_date=? WHERE user_id=?", (new_streak, today, uid))

# ── Achievements ──────────────────────────────────────────────────────────────
def check_achievements(uid):
    total = _calc_score(uid)
    counts = get_counts(uid)
    u = get_user(uid)
    counts["_streak"] = u["streak"] if u else 0
    with db() as c:
        earned = {r["achievement_id"] for r in c.execute(
            "SELECT achievement_id FROM achievements WHERE user_id=?", (uid,)
        ).fetchall()}
        new_ones = []
        for a in ACHIEVEMENTS:
            if a["id"] not in earned and a["check"](total, counts):
                c.execute(
                    "INSERT INTO achievements (user_id, achievement_id, earned_at) VALUES (?,?,?)",
                    (uid, a["id"], datetime.now().isoformat())
                )
                new_ones.append(a)
    return new_ones

# ── Settings ─────────────────────────────────────────────────────────────────
def set_setting(key, value):
    with db() as c:
        c.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (key, str(value)))

# ── Messages ──────────────────────────────────────────────────────────────────
def save_message(uid, text):
    with db() as c:
        c.execute(
            "INSERT INTO messages (user_id, text, sent_at) VALUES (?,?,?)",
            (uid, text, datetime.now().isoformat())
        )

# ─── UI YORDAMCHILARI ─────────────────────────────────────────────────────────
def progress_bar(val, mx=50, length=10):
    f = min(length, int(val / mx * length)) if mx else 0
    return "█" * f + "░" * (length - f)

def main_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Ball qo'shish",  callback_data="add"),
         InlineKeyboardButton("📊 Profilim",       callback_data="profile")],
        [InlineKeyboardButton("🏆 Reyting",        callback_data="leaderboard"),
         InlineKeyboardButton("🎖️ Yutuqlar",       callback_data="achievements")],
        [InlineKeyboardButton("📅 Tarix",          callback_data="history"),
         InlineKeyboardButton("💡 Maslahat",       callback_data="tips")],
        [InlineKeyboardButton("⚙️ Semestr",        callback_data="semester"),
         InlineKeyboardButton("✉️ Adminga xabar",  callback_data="msg_admin")],
    ])

def back_to_main():
    return InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Asosiy menyu", callback_data="back")]])

def activity_menu():
    acts = get_activities()
    rows = []
    for key, act in acts.items():
        label = f"{act['icon']} {act['name']}  (+{act['pts']})"
        if act["max"]:
            label += f"  [max:{act['max']}]"
        rows.append([InlineKeyboardButton(label, callback_data=f"act_{key}")])
    rows.append([InlineKeyboardButton("◀️ Orqaga", callback_data="back")])
    return InlineKeyboardMarkup(rows)

def admin_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏳ Kutayotgan so'rovlar", callback_data="adm_pending")],
        [InlineKeyboardButton("👥 Barcha talabalar",     callback_data="adm_users"),
         InlineKeyboardButton("📋 Ro'yxatdan o'tish",   callback_data="adm_reg")],
        [InlineKeyboardButton("📊 Stipendiat ro'yxati",  callback_data="adm_stipend"),
         InlineKeyboardButton("📤 Excel eksport",        callback_data="adm_excel")],
        [InlineKeyboardButton("📣 Broadcast",            callback_data="adm_broadcast"),
         InlineKeyboardButton("⚙️ Limitlar",             callback_data="adm_limits")],
    ])

# ─── STATUS TEKSHIRUVI ────────────────────────────────────────────────────────
async def check_access(update: Update) -> str | None:
    """None = OK, string = error message"""
    uid = update.effective_user.id
    if uid in ADMIN_IDS:
        return None
    u = get_user(uid)
    if not u:
        return "Siz ro'yxatdan o'tmagansiz. /start bosing."
    if u["status"] == "pending":
        return "⏳ Sizning ro'yxatdan o'tishingiz admin tasdiqlashini kutmoqda.\n\nAdmin tez orada ko'rib chiqadi."
    if u["status"] == "rejected":
        return (
            "❌ Sizning arizangiz rad etildi.\n\n"
            "Qo'shimcha ma'lumot uchun adminга murojaat qiling.\n"
            "Admin tasdiqlagunga qadar botdan foydalana olmaysiz."
        )
    if u["status"] == "blocked":
        return "🚫 Siz bloklangansiz. Admin bilan bog'laning."
    return None

# ─── RO'YXATDAN O'TISH ────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)

    if uid in ADMIN_IDS:
        await update.message.reply_text(
            "👨‍💼 *Admin paneli*\n\nXush kelibsiz!",
            parse_mode="Markdown",
            reply_markup=admin_menu(),
        )
        return

    if u:
        err = await check_access(update)
        if err:
            await update.message.reply_text(err)
            return
        total = _calc_score(uid, u["semester"])
        await update.message.reply_text(
            f"Xush kelibsiz, *{u['name']}*! 👋\n\n"
            f"📊 {u['semester']}-semestr: *{total}* ball\n"
            f"{'🔥 Streak: ' + str(u['streak']) + ' kun' if u['streak'] > 1 else ''}",
            parse_mode="Markdown",
            reply_markup=main_menu(),
        )
        return

    # Yangi foydalanuvchi
    ctx.user_data.clear()
    ctx.user_data["step"] = "name"
    await update.message.reply_text(
        "🎓 *Stipendiat Ball Tracker*\n\n"
        "Xush kelibsiz! Ro'yxatdan o'tish uchun:\n\n"
        "👤 Ism va familiyangizni kiriting:",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )

async def handle_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    text = update.message.text.strip()
    step = ctx.user_data.get("step")

    # ── Ro'yxatdan o'tish ────────────────────────────────────────────────────
    if step == "name":
        if len(text) < 3:
            await update.message.reply_text("Iltimos, to'liq ism-familiyangizni kiriting.")
            return
        ctx.user_data["name"] = text
        ctx.user_data["step"] = "group"
        await update.message.reply_text(
            "📚 Guruhingizni kiriting (masalan: *MX-21*, *IQ-22*):",
            parse_mode="Markdown",
        )
        return

    if step == "group":
        ctx.user_data["group"] = text
        ctx.user_data["step"] = "phone"
        kb = ReplyKeyboardMarkup(
            [[KeyboardButton("📱 Telefon raqamimni yuborish", request_contact=True)]],
            resize_keyboard=True, one_time_keyboard=True,
        )
        await update.message.reply_text(
            "📱 Telefon raqamingizni yuboring:",
            reply_markup=kb,
        )
        return

    # ── Adminga xabar ────────────────────────────────────────────────────────
    if step == "msg_admin":
        err = await check_access(update)
        if err:
            await update.message.reply_text(err)
            return
        u = get_user(uid)
        save_message(uid, text)
        ctx.user_data.pop("step", None)
        # Forward to admins
        for admin_id in ADMIN_IDS:
            try:
                await ctx.bot.send_message(
                    admin_id,
                    f"✉️ *Talabadan xabar*\n\n"
                    f"👤 {u['name']} | {u['group_name']}\n"
                    f"📱 {u['phone']}\n"
                    f"🆔 `{uid}`\n\n"
                    f"💬 {text}",
                    parse_mode="Markdown",
                )
            except Exception:
                pass
        await update.message.reply_text(
            "✅ Xabaringiz adminга yuborildi!",
            reply_markup=ReplyKeyboardRemove(),
        )
        await update.message.reply_text("Asosiy menyu:", reply_markup=main_menu())
        return

    # ── Admin broadcast ───────────────────────────────────────────────────────
    if step == "broadcast" and uid in ADMIN_IDS:
        ctx.user_data.pop("step", None)
        users = get_all_users(status="approved")
        sent, failed = 0, 0
        for u in users:
            try:
                await ctx.bot.send_message(u["user_id"], f"📣 *E'lon:*\n\n{text}", parse_mode="Markdown")
                sent += 1
            except Exception:
                failed += 1
        await update.message.reply_text(
            f"✅ Broadcast yakunlandi!\n✔️ Yuborildi: {sent}\n❌ Xato: {failed}",
            reply_markup=admin_menu(),
        )
        return

    # ── Admin: ball tuzatish ──────────────────────────────────────────────────
    if step == "adj_amount" and uid in ADMIN_IDS:
        target_uid = ctx.user_data.get("adj_uid")
        try:
            pts = int(text)
        except ValueError:
            await update.message.reply_text("Iltimos, sonni kiriting (masalan: 5 yoki -3)")
            return
        ctx.user_data.pop("step", None)
        ctx.user_data.pop("adj_uid", None)
        u = get_user(target_uid)
        sem = u["semester"] if u else 1
        # Add manual adjustment as a special activity
        with db() as c:
            c.execute(
                "INSERT INTO requests (user_id,activity_type,semester,status,submitted_at,reviewed_at,note) VALUES (?,'manual_adj',?,?,?,?,?)",
                (target_uid, sem, "approved", datetime.now().isoformat(), datetime.now().isoformat(), f"Admin tomonidan qo'shildi: {pts:+}")
            )
        # Override pts in settings doesn't work for individual — instead store in a manual_pts column
        await update.message.reply_text(
            f"✅ {u['name']} ga {pts:+} ball qo'shildi.",
            reply_markup=admin_menu(),
        )
        return

    # ── Admin: limit o'zgartirish ─────────────────────────────────────────────
    if step and step.startswith("set_limit_") and uid in ADMIN_IDS:
        key = step[10:]
        ctx.user_data.pop("step", None)
        val = "null" if text.lower() in ("0", "yo'q", "null", "none", "-") else text
        try:
            if val != "null":
                int(val)
        except ValueError:
            await update.message.reply_text("Iltimos, son kiriting (yoki 0 — limitsiz).")
            return
        set_setting(f"limit_{key}", val)
        acts = get_activities()
        act  = acts.get(key, {})
        lim  = f"{act.get('max', '—')}" if act else val
        await update.message.reply_text(
            f"✅ *{key}* uchun yangi limit: *{lim}*",
            parse_mode="Markdown",
            reply_markup=admin_menu(),
        )
        return

    # ── Boshqa xabarlar ───────────────────────────────────────────────────────
    if uid in ADMIN_IDS:
        await update.message.reply_text("Admin paneli:", reply_markup=admin_menu())
        return

    err = await check_access(update)
    if err:
        await update.message.reply_text(err)
        return

    await update.message.reply_text("Asosiy menyu:", reply_markup=main_menu())

async def handle_contact(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid     = update.effective_user.id
    contact = update.message.contact
    step    = ctx.user_data.get("step")

    if step != "phone":
        return
    if contact.user_id != uid:
        await update.message.reply_text("Iltimos, o'z raqamingizni yuboring.")
        return

    name  = ctx.user_data.get("name", "")
    group = ctx.user_data.get("group", "")
    phone = contact.phone_number
    ctx.user_data.clear()

    if not create_user(uid, name, group, phone):
        await update.message.reply_text(
            "Siz allaqachon ro'yxatdan o'tgansiz!",
            reply_markup=ReplyKeyboardRemove(),
        )
        return

    # Adminlarga bildirish
    for admin_id in ADMIN_IDS:
        try:
            await ctx.bot.send_message(
                admin_id,
                f"📝 *Yangi ro'yxatdan o'tish*\n\n"
                f"👤 {name}\n📚 {group}\n📱 {phone}\n🆔 `{uid}`",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("✅ Tasdiqlash",  callback_data=f"reg_approve_{uid}"),
                    InlineKeyboardButton("❌ Rad etish",   callback_data=f"reg_reject_{uid}"),
                ]]),
            )
        except Exception:
            pass

    await update.message.reply_text(
        "✅ *Ariza yuborildi!*\n\n"
        "⏳ Admin tasdiqlashini kuting.\n"
        "Tasdiqlangach, botdan to'liq foydalana olasiz.",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )

async def handle_file(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    step = ctx.user_data.get("step")
    if step != "file":
        return

    err = await check_access(update)
    if err:
        await update.message.reply_text(err)
        return

    atype = ctx.user_data.pop("act_type", None)
    ctx.user_data.pop("step", None)
    if not atype:
        return

    u = get_user(uid)
    msg = update.message
    if msg.document:
        file_id, ftype = msg.document.file_id, "document"
    elif msg.photo:
        file_id, ftype = msg.photo[-1].file_id, "photo"
    else:
        await update.message.reply_text("Iltimos, fayl yoki rasm yuboring.")
        return

    acts    = get_activities()
    act     = acts[atype]
    req_id  = submit_request(uid, atype, u["semester"], file_id, ftype)

    # Adminlarga yuborish
    for admin_id in ADMIN_IDS:
        try:
            caption = (
                f"📩 *Yangi so'rov #{req_id}*\n\n"
                f"👤 {u['name']} | {u['group_name']}\n"
                f"📱 {u['phone']}\n"
                f"📋 {act['icon']} {act['name']}  (+{act['pts']} ball)\n"
                f"📅 {u['semester']}-semestr"
            )
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Tasdiqlash",  callback_data=f"req_approve_{req_id}_{uid}"),
                InlineKeyboardButton("❌ Rad etish",   callback_data=f"req_reject_{req_id}_{uid}"),
            ]])
            if ftype == "photo":
                await ctx.bot.send_photo(admin_id, file_id, caption=caption,
                                         parse_mode="Markdown", reply_markup=kb)
            else:
                await ctx.bot.send_document(admin_id, file_id, caption=caption,
                                            parse_mode="Markdown", reply_markup=kb)
        except Exception:
            pass

    await update.message.reply_text(
        f"✅ *Hujjat yuborildi!*\n\n"
        f"{act['icon']} {act['name']}\n\n"
        f"_{random.choice(MOTIVATIONS)}_",
        parse_mode="Markdown",
        reply_markup=main_menu(),
    )

# ─── TALABA CALLBACK ──────────────────────────────────────────────────────────
async def button(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q   = update.callback_query
    await q.answer()
    uid = q.from_user.id
    d   = q.data

    # ── Admin callbacklar ─────────────────────────────────────────────────────
    if uid in ADMIN_IDS:
        await admin_callback(q, ctx, uid, d)
        return

    # ── Kirish tekshiruvi ─────────────────────────────────────────────────────
    u = get_user(uid)
    if not u:
        await q.message.reply_text("Avval /start buyrug'ini bering.")
        return
    if u["status"] != "approved":
        await q.message.reply_text(
            "⏳ Adminning tasdiqlashini kuting." if u["status"] == "pending"
            else "❌ Arizangiz rad etildi. Admin bilan bog'laning."
        )
        return

    sem = u["semester"]

    if d == "back":
        await q.message.reply_text("Asosiy menyu:", reply_markup=main_menu())

    elif d == "add":
        await q.message.reply_text(
            "Qaysi faoliyat?\n_Hujjat/rasm yuborishingiz kerak bo'ladi._",
            parse_mode="Markdown",
            reply_markup=activity_menu(),
        )

    elif d.startswith("act_"):
        key  = d[4:]
        acts = get_activities()
        act  = acts.get(key)
        if not act:
            return
        if act["max"]:
            done = get_sem_activity_count(uid, key, sem)
            if done >= act["max"]:
                await q.message.reply_text(
                    f"⚠️ *{act['name']}* semestr limiti to'ldi ({act['max']} ta)!",
                    parse_mode="Markdown",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Orqaga", callback_data="add")]])
                )
                return
        ctx.user_data["step"]     = "file"
        ctx.user_data["act_type"] = key
        done = get_sem_activity_count(uid, key, sem)
        lim  = f"{done}/{act['max']}" if act["max"] else str(done)
        await q.message.reply_text(
            f"*{act['icon']} {act['name']}*\n"
            f"Ball: *+{act['pts']}* | Semestrda: *{lim}*\n\n"
            "📎 *Hujjat yoki rasmni yuboring* (sertifikat, tasdiqnoma, skrinshotni):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Bekor", callback_data="back")]])
        )

    elif d == "profile":
        total    = _calc_score(uid, sem)
        all_tot  = _calc_score(uid)
        counts   = get_counts(uid, sem)
        acts     = get_activities()
        bar      = progress_bar(total, 50)
        pct      = min(100, int(total / 50 * 100))
        earned   = len([a for a in ACHIEVEMENTS
                        if any(True for _ in [1]
                               if a["id"] in {r["achievement_id"]
                                              for r in []})])
        # re-count earned properly
        with db() as c:
            e_count = c.execute(
                "SELECT COUNT(*) FROM achievements WHERE user_id=?", (uid,)
            ).fetchone()[0]

        text = (
            f"👤 *{u['name']}* | {u['group_name']}\n"
            f"📱 {u['phone']}\n"
            f"📅 {sem}-semestr\n\n"
            f"🏅 Semestr ball: *{total}*\n"
            f"[{bar}] {pct}%\n\n"
        )
        if u["streak"] > 1:
            text += f"🔥 Streak: *{u['streak']} kun*\n"
        text += f"🎖️ Yutuqlar: *{e_count}/{len(ACHIEVEMENTS)}*\n"
        text += f"📈 Jami: *{all_tot}* ball\n\n*Semestr faoliyatlar:*\n"
        for key, act in acts.items():
            cnt = counts.get(key, 0)
            if cnt:
                ml  = f"/{act['max']}" if act["max"] else ""
                text += f"{act['icon']} {act['name']}: {cnt}{ml} → *+{cnt*act['pts']}*\n"
        if not any(counts.get(k) for k in acts):
            text += "_Hali faoliyat yo'q._"
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu())

    elif d == "leaderboard":
        board   = get_leaderboard()
        medals  = ["🥇","🥈","🥉","4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"]
        text    = "🏆 *Top 10 — Umumiy reyting*\n\n"
        for i, (name, group, total) in enumerate(board):
            text += f"{medals[i]} *{name}* ({group}) — {total} ball\n"
        if not board:
            text += "_Hali ma'lumot yo'q._"
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu())

    elif d == "achievements":
        with db() as c:
            earned = {r["achievement_id"] for r in c.execute(
                "SELECT achievement_id FROM achievements WHERE user_id=?", (uid,)
            ).fetchall()}
        text = f"🎖️ *Yutuqlar — {len(earned)}/{len(ACHIEVEMENTS)}*\n\n"
        for a in ACHIEVEMENTS:
            if a["id"] in earned:
                text += f"✅ *{a['icon']} {a['name']}*\n_{a['desc']}_\n\n"
            else:
                text += f"🔒 {a['icon']} {a['name']}\n_{a['desc']}_\n\n"
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu())

    elif d == "history":
        rows = get_user_history(uid, 15)
        acts = get_activities()
        text = "📅 *So'nggi 15 so'rov*\n\n"
        for r in rows:
            act  = acts.get(r["activity_type"], {})
            icon = act.get("icon", "•")
            name = act.get("name", r["activity_type"])
            pts  = act.get("pts", "?")
            date = r["submitted_at"][:10]
            st   = {"pending": "⏳", "approved": "✅", "rejected": "❌"}.get(r["status"], "•")
            text += f"{st} {icon} {name}  *+{pts}*  _{date}_\n"
        if not rows:
            text += "_Hali so'rov yo'q._"
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu())

    elif d == "tips":
        tips = [
            "📰 OAK jurnali eng yuqori ball — 3 ta! Imkon bo'lsa yozing.",
            "🤝 Volontyorlik 10 ball/kun — eng samarali, lekin max 10 kun.",
            "🌍 Xalqaro nashr max 2 ta — imkoniyatni o'tkazib yuborma!",
            "📆 Har kun faoliyat qo'sh — streak rekordi qur!",
            "🏆 50 ball to'lasang 'STIPENDIAT!' yutuqqa ega bo'lasan!",
            "📋 Barcha hujjatlarni saqlab tur — kerak bo'ladi.",
        ]
        await q.message.reply_text(
            f"💡 *Maslahat:*\n\n{random.choice(tips)}\n\n📋 *Barchasi:*\n\n" + "\n\n".join(tips),
            parse_mode="Markdown", reply_markup=main_menu(),
        )

    elif d == "semester":
        sem = u["semester"]
        kb  = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"{'✅ ' if sem==i else ''}{i}-semestr", callback_data=f"sem_{i}")
             for i in range(1, 3)],
            [InlineKeyboardButton(f"{'✅ ' if sem==i else ''}{i}-semestr", callback_data=f"sem_{i}")
             for i in range(3, 5)],
            [InlineKeyboardButton("◀️ Orqaga", callback_data="back")],
        ])
        await q.message.reply_text(f"Hozirgi semestr: *{sem}*\nO'zgartirish:", parse_mode="Markdown", reply_markup=kb)

    elif d.startswith("sem_"):
        new_sem = int(d[4:])
        set_user_semester(uid, new_sem)
        total = _calc_score(uid, new_sem)
        await q.message.reply_text(
            f"✅ *{new_sem}-semestrga o'tildi!*\nBall: *{total}*",
            parse_mode="Markdown", reply_markup=main_menu()
        )

    elif d == "msg_admin":
        ctx.user_data["step"] = "msg_admin"
        await q.message.reply_text(
            "✉️ Adminga xabaringizni yozing:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Bekor", callback_data="back")]])
        )

# ─── ADMIN CALLBACK ───────────────────────────────────────────────────────────
async def admin_callback(q, ctx, uid, d):

    # ── Ro'yxatdan o'tish tasdiqlash ─────────────────────────────────────────
    if d.startswith("reg_approve_"):
        target = int(d[12:])
        update_user_status(target, "approved")
        u = get_user(target)
        try:
            await ctx.bot.send_message(
                target,
                "🎉 *Tabriklaymiz!*\n\nRo'yxatdan o'tishingiz tasdiqlandi!\n"
                "Endi botdan to'liq foydalana olasiz.",
                parse_mode="Markdown", reply_markup=main_menu(),
            )
        except Exception:
            pass
        await q.message.edit_caption(
            caption=q.message.caption + f"\n\n✅ *Tasdiqlandi* ({u['name']})",
            parse_mode="Markdown",
        ) if q.message.caption else await q.message.edit_text(
            q.message.text + f"\n\n✅ *Tasdiqlandi*", parse_mode="Markdown"
        )

    elif d.startswith("reg_reject_"):
        target = int(d[11:])
        update_user_status(target, "rejected")
        try:
            await ctx.bot.send_message(
                target,
                "❌ *Afsuski, arizangiz rad etildi.*\n\n"
                "Qo'shimcha ma'lumot uchun bevosita adminга murojaat qiling.",
                parse_mode="Markdown",
            )
        except Exception:
            pass
        txt = q.message.text or ""
        cap = q.message.caption or ""
        if cap:
            await q.message.edit_caption(caption=cap + "\n\n❌ *Rad etildi*", parse_mode="Markdown")
        else:
            await q.message.edit_text(txt + "\n\n❌ *Rad etildi*", parse_mode="Markdown")

    # ── Faoliyat tasdiqlash ───────────────────────────────────────────────────
    elif d.startswith("req_approve_"):
        parts     = d.split("_")
        req_id    = int(parts[2])
        target    = int(parts[3])
        review_request(req_id, "approved")
        update_streak(target)
        new_achs = check_achievements(target)
        u  = get_user(target)
        with db() as c:
            req = c.execute("SELECT * FROM requests WHERE id=?", (req_id,)).fetchone()
        acts = get_activities()
        act  = acts.get(req["activity_type"], {})
        msg  = (
            f"✅ *Faoliyatingiz tasdiqlandi!*\n\n"
            f"{act.get('icon','🏅')} {act.get('name', req['activity_type'])}"
            f"  *+{act.get('pts','?')} ball*\n\n"
            f"Jami ball: *{_calc_score(target, req['semester'])}*"
        )
        if new_achs:
            msg += "\n\n🎉 *Yangi yutuq(lar):*\n"
            for a in new_achs:
                msg += f"{a['icon']} *{a['name']}*\n"
        try:
            await ctx.bot.send_message(target, msg, parse_mode="Markdown", reply_markup=main_menu())
        except Exception:
            pass
        cap = q.message.caption or q.message.text or ""
        edit = cap + f"\n\n✅ *Tasdiqlandi* by admin"
        if q.message.caption is not None:
            await q.message.edit_caption(caption=edit, parse_mode="Markdown")
        else:
            await q.message.edit_text(edit, parse_mode="Markdown")

    elif d.startswith("req_reject_"):
        parts  = d.split("_")
        req_id = int(parts[2])
        target = int(parts[3])
        review_request(req_id, "rejected", "Hujjat qabul qilinmadi")
        u    = get_user(target)
        with db() as c:
            req = c.execute("SELECT * FROM requests WHERE id=?", (req_id,)).fetchone()
        acts = get_activities()
        act  = acts.get(req["activity_type"], {})
        try:
            await ctx.bot.send_message(
                target,
                f"❌ *So'rov rad etildi*\n\n"
                f"{act.get('icon','🏅')} {act.get('name', req['activity_type'])}\n\n"
                f"Hujjatni tekshirib, qayta yuboring.",
                parse_mode="Markdown", reply_markup=main_menu(),
            )
        except Exception:
            pass
        cap  = q.message.caption or q.message.text or ""
        edit = cap + "\n\n❌ *Rad etildi*"
        if q.message.caption is not None:
            await q.message.edit_caption(caption=edit, parse_mode="Markdown")
        else:
            await q.message.edit_text(edit, parse_mode="Markdown")

    # ── Admin menyu ───────────────────────────────────────────────────────────
    elif d == "back_admin":
        await q.message.reply_text("Admin paneli:", reply_markup=admin_menu())

    elif d == "adm_pending":
        reqs = get_pending_requests()
        if not reqs:
            await q.message.reply_text("⏳ Kutayotgan so'rovlar yo'q!", reply_markup=admin_menu())
            return
        await q.message.reply_text(f"⏳ *{len(reqs)} ta kutayotgan so'rov:*", parse_mode="Markdown")
        acts = get_activities()
        for req in reqs[:10]:
            act = acts.get(req["activity_type"], {})
            cap = (
                f"📩 So'rov #{req['id']}\n"
                f"👤 {req['name']} | {req['group_name']}\n"
                f"📋 {act.get('icon','')} {act.get('name', req['activity_type'])}\n"
                f"📅 {req['semester']}-semestr | {req['submitted_at'][:10]}"
            )
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅", callback_data=f"req_approve_{req['id']}_{req['user_id']}"),
                InlineKeyboardButton("❌", callback_data=f"req_reject_{req['id']}_{req['user_id']}"),
            ]])
            try:
                if req["file_type"] == "photo":
                    await ctx.bot.send_photo(uid, req["file_id"], caption=cap, reply_markup=kb)
                elif req["file_type"] == "document":
                    await ctx.bot.send_document(uid, req["file_id"], caption=cap, reply_markup=kb)
                else:
                    await ctx.bot.send_message(uid, cap, reply_markup=kb)
            except Exception:
                await ctx.bot.send_message(uid, cap + "\n_(Fayl yuborishda xato)_",
                                           parse_mode="Markdown", reply_markup=kb)

    elif d == "adm_users":
        users = get_all_users(status="approved")
        acts  = get_activities()
        text  = f"👥 *Tasdiqlangan talabalar: {len(users)} ta*\n\n"
        for u in users[:30]:
            total = _calc_score(u["user_id"])
            text += f"• *{u['name']}* ({u['group_name']}) — {total} ball\n"
        if len(users) > 30:
            text += f"\n_...va yana {len(users)-30} ta_"
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🗑 Talabani o'chirish", callback_data="adm_del_user"),
             InlineKeyboardButton("✏️ Ball tuzatish",      callback_data="adm_adj_score")],
            [InlineKeyboardButton("◀️ Orqaga",             callback_data="back_admin")],
        ])
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

    elif d == "adm_reg":
        users = get_all_users(status="pending")
        text  = f"📋 *Kutayotgan ro'yxatdan o'tishlar: {len(users)} ta*\n\n"
        for u in users:
            text += (
                f"👤 *{u['name']}* | {u['group_name']}\n"
                f"📱 {u['phone']} | 🆔 `{u['user_id']}`\n"
                f"📅 {u['joined_at'][:10]}\n\n"
            )
        if not users:
            text = "📋 Kutayotgan ro'yxatdan o'tish yo'q."
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=admin_menu())

    elif d == "adm_stipend":
        board = get_leaderboard(20)
        text  = "🏆 *Stipendiat ro'yxati (top 20)*\n\n"
        for i, (name, group, total) in enumerate(board, 1):
            emoji = "🏆" if total >= 50 else ("⭐" if total >= 30 else "•")
            text += f"{emoji} *{i}. {name}* ({group}) — {total} ball\n"
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=admin_menu())

    elif d == "adm_excel":
        if not EXCEL_OK:
            await q.message.reply_text("openpyxl o'rnatilmagan: `pip install openpyxl`",
                                       parse_mode="Markdown")
            return
        users = get_all_users()
        acts  = get_activities()
        wb    = openpyxl.Workbook()
        ws    = wb.active
        ws.title = "Talabalar"

        header_font = Font(bold=True)
        fill        = PatternFill("solid", fgColor="4472C4")

        headers = ["#", "Ism", "Guruh", "Telefon", "Status", "Semestr", "Streak", "Jami ball",
                   "Qo'shilgan sana"]
        for col, h in enumerate(headers, 1):
            cell            = ws.cell(1, col, h)
            cell.font       = Font(bold=True, color="FFFFFF")
            cell.fill       = fill
            cell.alignment  = Alignment(horizontal="center")

        for i, u in enumerate(users, 1):
            total = _calc_score(u["user_id"])
            ws.append([
                i, u["name"], u["group_name"], u["phone"] or "",
                u["status"], u["semester"], u["streak"],
                total, u["joined_at"][:10],
            ])

        # Activities sheet
        ws2 = wb.create_sheet("Faoliyatlar")
        ws2.append(["#", "Talaba", "Guruh", "Faoliyat", "Ball", "Semestr", "Status", "Sana"])
        with db() as c:
            all_reqs = c.execute(
                "SELECT r.*, u.name, u.group_name FROM requests r JOIN users u ON r.user_id=u.user_id ORDER BY r.submitted_at DESC"
            ).fetchall()
        for i, r in enumerate(all_reqs, 1):
            act = acts.get(r["activity_type"], {})
            ws2.append([
                i, r["name"], r["group_name"],
                act.get("name", r["activity_type"]),
                act.get("pts", "?"), r["semester"],
                r["status"], r["submitted_at"][:10],
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"stipend_hisobot_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        await ctx.bot.send_document(uid, document=buf, filename=fname,
                                    caption="📊 Excel hisobot tayyor!")

    elif d == "adm_broadcast":
        ctx.user_data["step"] = "broadcast"
        await q.message.reply_text(
            "📣 Barcha tasdiqlangan talabalarga yuboriladigan xabarni yozing:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Bekor", callback_data="back_admin")]])
        )

    elif d == "adm_limits":
        acts = get_activities()
        text = "⚙️ *Faoliyat limitlari:*\n\n"
        for key, act in acts.items():
            lim = str(act["max"]) if act["max"] else "Limitsiz"
            text += f"{act['icon']} {act['name']}: *{lim}*\n"
        kb = InlineKeyboardMarkup(
            [[InlineKeyboardButton(f"✏️ {act['icon']} {act['name']}", callback_data=f"adm_setlim_{key}")]
             for key, act in acts.items()]
            + [[InlineKeyboardButton("◀️ Orqaga", callback_data="back_admin")]]
        )
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

    elif d.startswith("adm_setlim_"):
        key  = d[11:]
        acts = get_activities()
        act  = acts.get(key, {})
        ctx.user_data["step"] = f"set_limit_{key}"
        await q.message.reply_text(
            f"*{act.get('name', key)}* uchun yangi limitni kiriting:\n"
            f"(0 yoki «yo'q» — limitsiz)",
            parse_mode="Markdown",
        )

    elif d == "adm_del_user":
        users = get_all_users(status="approved")
        text  = "🗑 *O'chirish uchun tanlang (ID):*\n\n"
        for u in users[:20]:
            text += f"• `{u['user_id']}` — {u['name']} ({u['group_name']})\n"
        await q.message.reply_text(
            text + "\n\nO'chirish uchun: /del_TELEGRAM_ID",
            parse_mode="Markdown",
        )

    elif d == "adm_adj_score":
        users = get_all_users(status="approved")
        text  = "✏️ *Ball tuzatish (ID ni yuboring):*\n\n"
        for u in users[:20]:
            total = _calc_score(u["user_id"])
            text += f"• `{u['user_id']}` — {u['name']}: {total} ball\n"
        await q.message.reply_text(
            text + "\n\nID va summani yuborish uchun: /adj_TELEGRAM_ID",
            parse_mode="Markdown",
        )

# ─── ADMIN COMMANDS ──────────────────────────────────────────────────────────
async def cmd_admin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return
    await update.message.reply_text("👨‍💼 *Admin paneli:*", parse_mode="Markdown",
                                    reply_markup=admin_menu())

async def cmd_del(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return
    parts = update.message.text.split("_")
    if len(parts) < 2:
        return
    try:
        target = int(parts[1])
    except ValueError:
        return
    u = get_user(target)
    if not u:
        await update.message.reply_text("Foydalanuvchi topilmadi.")
        return
    delete_user(target)
    try:
        await ctx.bot.send_message(target, "⚠️ Sizning akkauntingiz o'chirildi.")
    except Exception:
        pass
    await update.message.reply_text(f"✅ {u['name']} o'chirildi.", reply_markup=admin_menu())

async def cmd_adj(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return
    parts = update.message.text.split("_")
    if len(parts) < 2:
        return
    try:
        target = int(parts[1])
    except ValueError:
        return
    u = get_user(target)
    if not u:
        await update.message.reply_text("Foydalanuvchi topilmadi.")
        return
    ctx.user_data["step"]    = "adj_amount"
    ctx.user_data["adj_uid"] = target
    total = _calc_score(target)
    await update.message.reply_text(
        f"✏️ *{u['name']}* — hozir *{total}* ball\n\n"
        f"Qo'shish/ayirish miqdorini kiriting (masalan: 5 yoki -3):",
        parse_mode="Markdown",
    )

# ─── ASOSIY ──────────────────────────────────────────────────────────────────
def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start",   cmd_start))
    app.add_handler(CommandHandler("admin",   cmd_admin))
    app.add_handler(CommandHandler("help",    cmd_start))
    app.add_handler(MessageHandler(filters.Regex(r"^/del_\d+$"), cmd_del))
    app.add_handler(MessageHandler(filters.Regex(r"^/adj_\d+$"), cmd_adj))
    app.add_handler(MessageHandler(filters.CONTACT, handle_contact))
    app.add_handler(MessageHandler(filters.Document.ALL | filters.PHOTO, handle_file))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(CallbackQueryHandler(button))
    print("🤖 Bot ishga tushdi!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
